from openpyxl import *
from collections import OrderedDict
import openpyxl


def createColumnHeaders(headers_list, wb_name, ws_name, rowNumber=1,
                        overwrite=False, create_ws=False):
    try:
        headers_dictionary = {}
        if ".xlsx" not in wb_name:
            wb_name += '.xlsx'
        wb = openpyxl.load_workbook(wb_name)
        try:
            ws = wb.get_sheet_by_name(name=ws_name)
            if overwrite:
                for i in range(1, len(headers_list) + 1):
                    ws.cell(row=rowNumber, column=i, value=headers_list[i - 1])
                    headers_dictionary[headers_list[i - 1]] = i
                wb.save(wb_name)
                return headers_dictionary
            else:
                raise ValueError(
                    "You are trying to write headers to a worksheet that " +
                    "already exists but overwrite is set to False."
                )

        except KeyError:
            if(create_ws):
                ws = wb.create_sheet(title=ws_name)
                for i in range(1, len(headers) + 1):
                    ws.cell(row=rowNumber, column=i, value=headers_list[i - 1])
                    headers_dictionary[headers_list[i - 1]] = i
                wb.save(wb_name)
                return headers_dictionary
            else:
                raise ValueError(
                                "You are trying to write headers to a " +
                                "worksheet that does not exist and " +
                                "create_ws is set to False."
                )
    except FileNotFoundError:  # raise error if workbook does not exist
        raise ValueError(
                        "You are trying to write headers to a workbook" +
                        "that does not exist."
        )


def getColumnHeaders(rowNumber, wb_name, ws_name=None, key='string'):
    try:
        if ".xlsx" not in wb_name:
            wb_name = wb_name + '.xlsx'
        wb = openpyxl.load_workbook(wb_name)  # check if workbook exists

        headers = {}

        if ws_name is None:
            ws = wb.active
        else:
            ws = wb.get_sheet_by_name(name=ws_name)
        try:  # make sure worksheet is valid
            for i in range(1, ws.max_column + 1):
                cell = ws.cell(row=rowNumber, column=i)
                if cell.value is not None:
                    if type(key) == str:
                        headers[str(ws.cell(row=rowNumber, column=i).value)] = i  # noqa
                    else:
                        headers[i] = str(ws.cell(row=rowNumber, column=i).value)  # noqa
        except:
            raise ValueError(
                "An error occurred while attempting" +
                " to retrieve the headers from the excel file."
            )
        return headers

    except FileNotFoundError:  # raise error if workbook does not exist
        raise ValueError(
            "You are trying to read headers to a workbook that does not" +
            "exist."
        )


def createExampleExcelFile(wb_name):
    wb = Workbook()
    ws = wb.active
    headerNames = ["a", "b", "c"]
    if ".xlsx" not in wb_name:
        wb_name += ".xlsx"
    wb.save(wb_name)
    headers = createColumnHeaders(headers_list=headerNames, wb_name=wb_name,
                                  ws_name=ws.title, overwrite=True)


def exampleUsage():
    createExampleExcelFile(wb_name="example.xlsx")
    # createColumnHeaders(headers_list = ["i", "j", "k"], wb_name = "example",
    # ws_name = "Sheet", overwrite = True)
    # print(getColumnHeaders(rowNumber = 1, wb_name = "example.xlsx"))

if __name__ == "__main__":
    exampleUsage()
